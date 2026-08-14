"""Create one styled six-sheet dashboard workbook per reporting month."""

from __future__ import annotations

import json
import math
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_FILE = ROOT / "assets" / "data" / "dashboard.json"
OUTPUT_DIR = ROOT / "assets" / "downloads"

BLUE = "0072BC"
BLUE_DARK = "0B3754"
BLUE_PALE = "E8F2F8"
INK = "242A2E"
INK_MUTED = "5F6B72"
LINE = "E3E8EC"
WHITE = "FFFFFF"
SHEET_NAMES = [
    "Demographics",
    "Infrastructure",
    "NFI Needs",
    "Education & Catering",
    "Services",
    "Calendar",
]
CALENDAR_DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
CALENDAR_TIMES = ["08:00–10:00", "10:00–12:00", "14:00–16:00", "16:00–18:00"]


def is_present(value) -> bool:
    return bool(str(value or "").strip())


def sort_racs(record: dict):
    value = str(record.get("racId", ""))
    return (0, int(value)) if value.isdigit() else (1, value)


def clean(value):
    return "" if value is None else value


def iso_date(value):
    if not value:
        return ""
    return datetime.fromisoformat(value).date()


def calendar_cell(entries: list) -> str:
    return "\n".join(
        f"{CALENDAR_TIMES[index]} — {entry}"
        for index, entry in enumerate(entries or [])
        if is_present(entry)
    )


def estimated_lines(value, characters_per_line: int) -> int:
    lines = str(value or "").splitlines() or [""]
    return sum(max(1, math.ceil(len(line) / characters_per_line)) for line in lines)


def sheet_data(payload: dict, month: str) -> list[dict]:
    records = sorted((record for record in payload["records"] if record["month"] == month), key=sort_racs)
    demographics = sorted(
        (record for record in payload["demographicsRecords"] if record["month"] == month),
        key=sort_racs,
    )

    needs_labels = list(records[0].get("needs", {})) if records else []
    needs_labels = [
        label for label in needs_labels
        if any(is_present(record.get("needs", {}).get(label)) for record in records)
    ]
    service_labels = list(records[0].get("services", {})) if records else []
    service_labels = [
        label for label in service_labels
        if any(is_present(record.get("services", {}).get(label)) for record in records)
    ]

    calendar_records = [
        record for record in records
        if any(any(is_present(entry) for entry in record.get("calendar", {}).get(day, [])) for day in CALENDAR_DAYS)
    ]

    return [
        {
            "name": "Demographics",
            "headers": ["RAC ID", "Raion", "Address", "Capacity", "Hosted", "0-17", "18-59", "60+", "PwD", "Primary source", "Data date"],
            "rows": [
                [
                    record["racId"], clean(record.get("raion")), clean(record.get("address")),
                    clean(record.get("capacity")), clean(record.get("hosted")),
                    clean(record.get("demographicProfile", {}).get("0-17 years")),
                    clean(record.get("demographicProfile", {}).get("18-59 years")),
                    clean(record.get("demographicProfile", {}).get("60+ years")),
                    clean(record.get("pwd")), clean(record.get("demographicSource")),
                    iso_date(record.get("demographicDate")),
                ]
                for record in demographics
            ],
            "numeric": {4, 5, 6, 7, 8, 9},
            "dates": {11},
            "wrap": {3},
        },
        {
            "name": "Infrastructure",
            "headers": ["RAC ID", "Raion", "Room type", "Toilet type", "Showers separate", "Separated by gender", "Accessible sanitation", "Child-friendly space", "Location", "Accessible entrance", "Adaptable for PwD"],
            "rows": [
                [
                    record["racId"], clean(record.get("raion")), clean(record.get("roomType")),
                    clean(record.get("toiletType")), clean(record.get("showersSeparate")),
                    clean(record.get("genderSeparatedSanitation")), clean(record.get("accessibleSanitation")),
                    clean(record.get("childFriendlySpace")), clean(record.get("childFriendlyLocation")),
                    clean(record.get("accessibleEntrance")), clean(record.get("adaptablePwd")),
                ]
                for record in records
            ],
            "numeric": set(),
            "dates": set(),
            "wrap": {3, 4},
        },
        {
            "name": "NFI Needs",
            "headers": ["RAC ID", "Raion", *needs_labels],
            "rows": [
                [record["racId"], clean(record.get("raion")), *[clean(record.get("needs", {}).get(label)) for label in needs_labels]]
                for record in records
            ],
            "numeric": set(),
            "dates": set(),
            "wrap": set(range(3, len(needs_labels) + 3)),
        },
        {
            "name": "Education & Catering",
            "headers": ["RAC ID", "Raion", "Hosted", "Children", "3-6", "7-11", "12-17", "Education access", "Attending", "In person", "Online", "Meal provider", "Food service", "Meals/day", "Satisfaction"],
            "rows": [
                [
                    record["racId"], clean(record.get("raion")), clean(record.get("hosted")),
                    clean(record.get("children")), clean(record.get("educationAges", {}).get("3-6 years")),
                    clean(record.get("educationAges", {}).get("7-11 years")),
                    clean(record.get("educationAges", {}).get("12-17 years")),
                    clean(record.get("educationAccess")), clean(record.get("schoolAttendance")),
                    clean(record.get("schoolInPerson")), clean(record.get("schoolOnline")),
                    clean(record.get("mealProvider")), clean(record.get("foodService")),
                    clean(record.get("mealsPerDay")), clean(record.get("cateringSatisfaction")),
                ]
                for record in records
            ],
            "numeric": {3, 4, 5, 6, 7, 9, 10, 11},
            "dates": set(),
            "wrap": {8, 12, 13, 15},
        },
        {
            "name": "Services",
            "headers": ["RAC ID", "Raion", *service_labels],
            "rows": [
                [record["racId"], clean(record.get("raion")), *[clean(record.get("services", {}).get(label)) for label in service_labels]]
                for record in records
            ],
            "numeric": set(),
            "dates": set(),
            "wrap": set(range(3, len(service_labels) + 3)),
        },
        {
            "name": "Calendar",
            "headers": ["RAC ID", "Raion", *CALENDAR_DAYS],
            "rows": [
                [
                    record["racId"], clean(record.get("raion")),
                    *[calendar_cell(record.get("calendar", {}).get(day, [])) for day in CALENDAR_DAYS],
                ]
                for record in calendar_records
            ],
            "numeric": set(),
            "dates": set(),
            "wrap": set(range(3, 10)),
        },
    ]


def column_width(header: str, values: list, column_number: int, calendar: bool) -> float:
    content_width = max([len(str(header)), *[max((len(line) for line in str(value or "").splitlines()), default=0) for value in values]])
    if column_number == 1:
        return 11
    if header == "Address":
        return min(max(content_width + 2, 24), 44)
    if calendar and column_number >= 3:
        return 42
    return min(max(content_width + 2, 12), 28)


def style_sheet(ws, spec: dict, month: str, meta: dict) -> None:
    headers = spec["headers"]
    rows = spec["rows"]
    last_column = get_column_letter(len(headers))
    month_label = datetime.strptime(month, "%Y-%m").strftime("%B %Y")
    title = f"RAC Monitoring Visits — {month_label}"
    subtitle = (
        f"Reporting month: {month}  |  Collection window: {meta['from']} to {meta['to']}  |  "
        "Sources: MLSP and ACTED"
    )

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{last_column}{max(4, len(rows) + 4)}"
    ws.merge_cells(f"A1:{last_column}1")
    ws.merge_cells(f"A2:{last_column}2")
    ws["A1"] = title
    ws["A2"] = subtitle
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE_DARK)
    ws["A1"].font = Font(name="Aptos Display", size=16, bold=True, color=WHITE)
    ws["A1"].alignment = Alignment(vertical="center")
    ws["A2"].fill = PatternFill("solid", fgColor=BLUE_PALE)
    ws["A2"].font = Font(name="Aptos", size=9, color=BLUE_DARK)
    ws["A2"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 29
    ws.row_dimensions[2].height = 23
    ws.row_dimensions[3].height = 8

    thin_line = Side(style="thin", color=LINE)
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(4, column, header)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(name="Aptos", size=9, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin_line)
    ws.row_dimensions[4].height = 31

    if rows:
        for row_index, values in enumerate(rows, start=5):
            for column, value in enumerate(values, start=1):
                cell = ws.cell(row_index, column, value)
                cell.font = Font(name="Aptos", size=9, color=INK)
                cell.alignment = Alignment(
                    vertical="top" if column in spec["wrap"] else "center",
                    wrap_text=column in spec["wrap"],
                )
                cell.border = Border(bottom=thin_line)
                if row_index % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F7F9FA")
                if column in spec["numeric"]:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif column in spec["dates"] and value:
                    cell.number_format = "yyyy-mm-dd"
            calendar_sheet = spec["name"] == "Calendar"
            line_count = max(
                (
                    estimated_lines(value, 42 if calendar_sheet else 26)
                    if column in spec["wrap"] else 1
                )
                for column, value in enumerate(values, start=1)
            )
            maximum_height = 390 if calendar_sheet else 66
            ws.row_dimensions[row_index].height = min(max(19, 12 * line_count), maximum_height)

    else:
        ws["A5"] = "No data reported for this month."
        ws["A5"].font = Font(name="Aptos", size=10, italic=True, color=INK_MUTED)

    for column, header in enumerate(headers, start=1):
        values = [row[column - 1] for row in rows]
        ws.column_dimensions[get_column_letter(column)].width = column_width(
            header, values, column, spec["name"] == "Calendar"
        )

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:4"
    ws.oddFooter.center.text = "RAC Monitoring Visits | MLSP and ACTED"
    ws.oddFooter.right.text = "Page &P of &N"


def build_monthly_workbooks(payload: dict) -> list[Path]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    month_meta = {month["id"]: month for month in payload["months"]}
    outputs = []

    for month in month_meta:
        workbook = Workbook()
        default_sheet = workbook.active
        specs = sheet_data(payload, month)

        for index, spec in enumerate(specs):
            worksheet = default_sheet if index == 0 else workbook.create_sheet()
            worksheet.title = spec["name"]
            worksheet.sheet_properties.tabColor = BLUE if index < 5 else BLUE_DARK
            style_sheet(worksheet, spec, month, month_meta[month])

        workbook.active = 0
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        destination = OUTPUT_DIR / f"rac-monitoring-visits-{month}.xlsx"
        workbook.save(destination)
        workbook.close()
        outputs.append(destination)

    return outputs


def validate_workbooks(payload: dict, outputs: list[Path]) -> None:
    expected_months = [month["id"] for month in payload["months"]]
    assert len(outputs) == len(expected_months)
    for month, path in zip(expected_months, outputs):
        workbook = load_workbook(path, read_only=False, data_only=False)
        try:
            assert workbook.sheetnames == SHEET_NAMES
            assert all(workbook[sheet]["A1"].value.startswith("RAC Monitoring Visits") for sheet in SHEET_NAMES)
            assert all(f"Reporting month: {month}" in workbook[sheet]["A2"].value for sheet in SHEET_NAMES)
            assert workbook["Demographics"]["A4"].value == "RAC ID"
            assert workbook["Calendar"]["C4"].value == "Monday"
            assert all(not workbook[sheet].sheet_view.showGridLines for sheet in SHEET_NAMES)
        finally:
            workbook.close()


def main() -> None:
    payload = json.loads(DATA_FILE.read_text(encoding="utf-8"))
    outputs = build_monthly_workbooks(payload)
    validate_workbooks(payload, outputs)
    print(f"Wrote and validated {len(outputs)} monthly workbooks in {OUTPUT_DIR.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
