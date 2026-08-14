"""Build the browser-ready RAC monitoring dataset.

ACTED monitoring data remain the source for round coverage and all thematic tabs.
For demographic measures, the latest MLSP snapshot within each calendar month
is preferred by RAC ID; ACTED fills RACs or months absent from that snapshot.
Capacity is always carried forward from the latest MLSP value available for the
same RAC on or before the selected reporting month.
RAC coordinates come from rac_map.xlsx and are matched only by RAC ID.
"""

from __future__ import annotations

import json
import math
import os
import shutil
import subprocess
import tempfile
from datetime import date, datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
KOBO_FILE = ROOT / "data" / "kobo_data.xlsx"
MLSP_FILE = ROOT / "data" / "rac_demographics.xlsx"
MAP_FILE = ROOT / "data" / "rac_map.xlsx"
TARGET = ROOT / "assets" / "data" / "dashboard.json"

# ACTED used 573 once for the Florilor site that MLSP and the map identify as 574.
# Keep the source record, but inherit capacity from the matching MLSP RAC history.
CAPACITY_RAC_ALIASES = {"573": "574"}


def report_month(value: datetime) -> str:
    """Assign ACTED visits on days 1-7 to the preceding reporting month."""
    year, month = value.year, value.month
    if value.day <= 7:
        month -= 1
        if month == 0:
            year -= 1
            month = 12
    return f"{year:04d}-{month:02d}"


def iso(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    return None


def number(value):
    if value in (None, ""):
        return 0
    try:
        parsed = float(value)
        if not math.isfinite(parsed):
            return 0
        return int(parsed) if parsed.is_integer() else parsed
    except (TypeError, ValueError):
        return 0


def optional_number(value):
    if value in (None, ""):
        return None
    try:
        parsed = float(value)
        if not math.isfinite(parsed):
            return None
        return int(parsed) if parsed.is_integer() else parsed
    except (TypeError, ValueError):
        return None


def text(value):
    return "" if value is None else str(value).strip()


def normalized_header(value):
    return " ".join(text(value).split()).lower()


def rac_id(value):
    if value in (None, ""):
        return ""
    try:
        parsed = float(value)
        return str(int(parsed)) if parsed.is_integer() else str(parsed)
    except (TypeError, ValueError):
        return text(value)


def _powershell_literal(value: Path) -> str:
    return "'" + str(value).replace("'", "''") + "'"


def readable_workbook(path: Path, temp_root: Path) -> Path:
    """Copy a workbook locally, including OneDrive files held as reparse points."""
    destination = temp_root / path.name
    try:
        shutil.copyfile(path, destination)
    except PermissionError:
        if os.name != "nt":
            raise
        command = (
            f"Copy-Item -LiteralPath {_powershell_literal(path)} "
            f"-Destination {_powershell_literal(destination)} -Force"
        )
        subprocess.run(
            ["powershell.exe", "-NoProfile", "-NonInteractive", "-Command", command],
            check=True,
            capture_output=True,
            text=True,
        )
    return destination


def load_mlsp_snapshots(path: Path) -> dict:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows)
        index = {normalized_header(header): position for position, header in enumerate(headers)}
        by_month = {}

        for row in rows:
            snapshot_date = row[index["date"]]
            uid = rac_id(row[index["uid"]])
            if not isinstance(snapshot_date, datetime) or not uid:
                continue
            month = snapshot_date.strftime("%Y-%m")
            current = by_month.get(month)
            if current is None or snapshot_date > current["date"]:
                by_month[month] = {"date": snapshot_date, "rows": {}}
            if snapshot_date == by_month[month]["date"]:
                by_month[month]["rows"][uid] = {
                    "capacity": optional_number(row[index["capacity"]]),
                    "hosted": optional_number(row[index["currently occupied"]]),
                    "age0to2": optional_number(row[index["children 0-2 years old"]]),
                    "age2to18": optional_number(row[index["children 2-18 years old"]]),
                    "age60plus": optional_number(row[index["individuals 60+"]]),
                    "pwd": optional_number(row[index["pwd"]]),
                    "wheelchair": optional_number(row[index["individuals with locomotor disabilities (wheelchair)"]]),
                    "bedridden": optional_number(row[index["bed-ridden individuals"]]),
                    "childrenWithDisability": optional_number(row[index["children with disability"]]),
                    "pregnantWomen": optional_number(row[index["pregnant women"]]),
                    "address": text(row[index["address"]]),
                }
        return by_month
    finally:
        workbook.close()


def load_locations(path: Path) -> list:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows)
        index = {normalized_header(header): position for position, header in enumerate(headers)}
        locations = []
        for row in rows:
            uid = rac_id(row[index["rac_id"]])
            latitude = optional_number(row[index["lat"]])
            longitude = optional_number(row[index["lon"]])
            if uid and latitude is not None and longitude is not None:
                locations.append({"racId": uid, "latitude": latitude, "longitude": longitude})
        return sorted(locations, key=lambda item: int(item["racId"]) if item["racId"].isdigit() else item["racId"])
    finally:
        workbook.close()


def kobo_profile(record: dict) -> dict:
    return {
        "0-17 years": number(record.get("children")),
        "18-59 years": number(record.get("age18to59")),
        "60+ years": number(record.get("age60plus")),
    }


def mlsp_demographic_record(
    base: dict,
    values: dict,
    month: str,
    snapshot_date: datetime,
    carried_capacity,
) -> dict:
    record = dict(base)
    hosted = values["hosted"] if values["hosted"] is not None else number(base.get("hosted"))
    age0to2 = values["age0to2"] if values["age0to2"] is not None else 0
    age2to18 = values["age2to18"] if values["age2to18"] is not None else 0
    age60plus = values["age60plus"] if values["age60plus"] is not None else number(base.get("age60plus"))
    children = age0to2 + age2to18
    age18to59 = max(hosted - children - age60plus, 0)
    record.update(
        {
            "month": month,
            "capacity": values["capacity"] if values["capacity"] is not None else carried_capacity,
            "hosted": hosted,
            "children": children,
            "age18to59": age18to59,
            "age60plus": age60plus,
            "pwd": values["pwd"] if values["pwd"] is not None else number(base.get("pwd")),
            "demographicProfile": {
                "0-17 years": children,
                "18-59 years": age18to59,
                "60+ years": age60plus,
            },
            "demographicSource": "MLSP",
            "demographicDate": iso(snapshot_date),
            "mlspDetails": {
                "0-2 years": age0to2,
                "2-18 years": age2to18,
                "wheelchair": values["wheelchair"],
                "bedridden": values["bedridden"],
                "childrenWithDisability": values["childrenWithDisability"],
                "pregnantWomen": values["pregnantWomen"],
            },
        }
    )
    if not record.get("address"):
        record["address"] = values["address"]
    return record


def main() -> None:
    with tempfile.TemporaryDirectory(prefix="rac-dashboard-") as temp_name:
        temp_root = Path(temp_name)
        mlsp_path = readable_workbook(MLSP_FILE, temp_root)
        map_path = readable_workbook(MAP_FILE, temp_root)
        mlsp_snapshots = load_mlsp_snapshots(mlsp_path)
        locations = load_locations(map_path)

    workbook = openpyxl.load_workbook(KOBO_FILE, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    index = {header: position for position, header in enumerate(headers) if header}

    ventilation_start = index["**Do these spaces have good ventilation?**"]
    privacy_start = index["**Do these spaces have adequate privacy?**"]
    security_start = index["**Do these spaces have adequate security? **"]
    pwd_spaces_start = index["**Are some of these spaces dedicated for PwD?**"]
    calendar_days = {
        "Monday": index["** Monday**"],
        "Tuesday": index["** Tuesday**"],
        "Wednesday": index["** Wednesday**"],
        "Thursday": index["** Thursday**"],
        "Friday": index["** Friday**"],
        "Saturday": index["** Saturday**"],
        "Sunday": index["** Sunday**"],
    }

    def value(row, header, default=""):
        position = index.get(header)
        return default if position is None else row[position]

    def column(row, position, default=""):
        return default if position < 0 or position >= len(row) else row[position]

    deduplicated = {}
    for row in rows:
        visit = value(row, "Date of visit", None)
        uid = rac_id(value(row, "RAC ID", None))
        if not isinstance(visit, datetime) or not uid:
            continue
        key = (report_month(visit), uid)
        end = value(row, "end", visit)
        if key not in deduplicated or end > deduplicated[key][0]:
            deduplicated[key] = (end, row)

    records = []
    for (month, uid), (_, row) in sorted(deduplicated.items()):
        female = [
            number(value(row, '<span style="display:none">two_female</span>')),
            number(value(row, '<span style="display:none">seventeen_female</span>')),
            number(value(row, '<span style="display:none">sixtyfour_female</span>')),
            number(value(row, '<span style="display:none">eighteen_female</span>')),
            number(value(row, '<span style="display:none">sixtyfive_female</span>')),
        ]
        male = [
            number(value(row, '<span style="display:none">two_male</span>')),
            number(value(row, '<span style="display:none">seventeen_male</span>')),
            number(value(row, '<span style="display:none">sixtyfour_male</span>')),
            number(value(row, '<span style="display:none">eighteen_male</span>')),
            number(value(row, '<span style="display:none">sixtyfive_male</span>')),
        ]
        other = [
            number(value(row, '<span style="display:none">two_dontId</span>')),
            number(value(row, '<span style="display:none">seventeen_dontId</span>')),
            number(value(row, '<span style="display:none">sixtyfour_dontId</span>')),
            number(value(row, '<span style="display:none">eighteen_dontId</span>')),
            number(value(row, '<span style="display:none">sixtyfive_dontId</span>')),
        ]
        ages = [female[i] + male[i] + other[i] for i in range(5)]
        record = {
            "month": month,
            "visitDate": iso(value(row, "Date of visit", None)),
            "racId": uid,
            "raion": text(value(row, "Raion")),
            "city": text(value(row, "City/municipality")),
            "address": text(value(row, "Centre's address")),
            "capacity": None,
            "hosted": number(value(row, "How many refugees are currently hosted by the centre?")),
            "female": sum(female),
            "male": sum(male),
            "otherGender": sum(other),
            "age0to4": ages[0],
            "age5to11": ages[1],
            "age12to17": ages[2],
            "age18to59": ages[3],
            "age60plus": ages[4],
            "ageGender": {
                "0-4 years": {"female": female[0], "male": male[0], "other": other[0]},
                "5-11 years": {"female": female[1], "male": male[1], "other": other[1]},
                "12-17 years": {"female": female[2], "male": male[2], "other": other[2]},
                "18-59 years": {"female": female[3], "male": male[3], "other": other[3]},
                "60+ years": {"female": female[4], "male": male[4], "other": other[4]},
            },
            "children": sum(ages[:3]),
            "pwd": number(value(row, "How many people with disabilities are in the Center in total?")),
            "educationAccess": text(value(row, "Do children hosted by the Centre have access to education services (kindergarten, school, summer camp)?")),
            "schoolAttendance": number(value(row, '<span style="display:none">total-In total</span>')),
            "schoolInPerson": number(value(row, '<span style="display:none">person-In person</span>')),
            "schoolOnline": number(value(row, '<span style="display:none">online-Online</span>')),
            "educationAges": {
                "3-6 years": number(value(row, "tot_ch3-6")),
                "7-11 years": number(value(row, "tot_ch7-11")),
                "12-17 years": number(value(row, "tot_ch12-17")),
            },
            "roomType": text(value(row, "Select the type of rooms")),
            "toiletType": text(value(row, "Select the type of toilets")),
            "showersSeparate": text(value(row, "Are the showers separate from the toilets?")),
            "genderSeparatedSanitation": text(value(row, "Are the showers and toilets separated by gender?")),
            "kitchen": text(value(row, "Is there a kitchen?")),
            "fireExtinguishers": text(value(row, "Are there fire extinguishers in the kitchen?")),
            "accessibleSanitation": text(value(row, "Are the showers and toilets equipped for people with disabilities?")),
            "diningRoom": text(value(row, "Is there a dining room?")),
            "childFriendlySpace": text(value(row, "Is there a specific space for children to play?")),
            "childFriendlyLocation": text(value(row, "Is the available space inside the building or outside?")),
            "ventilation": {
                "toilets": text(column(row, ventilation_start + 1)),
                "showers": text(column(row, ventilation_start + 8)),
                "kitchen": text(column(row, ventilation_start + 15)),
                "bedrooms": text(column(row, ventilation_start + 22)),
            },
            "privacy": {
                "bedrooms": text(column(row, privacy_start + 1)),
                "toilets": text(column(row, privacy_start + 9)),
                "showers": text(column(row, privacy_start + 17)),
            },
            "security": {
                "toilets": text(column(row, security_start + 1)),
                "showers": text(column(row, security_start + 9)),
                "kitchen": text(column(row, security_start + 17)),
                "bedrooms": text(column(row, security_start + 25)),
            },
            "adaptablePwd": text(value(row, "Is it possible to make any adaptation in the building to include spaces for PwD?")),
            "pwdSpaces": {
                "toilets": text(column(row, pwd_spaces_start + 1)),
                "showers": text(column(row, pwd_spaces_start + 2)),
                "kitchen": text(column(row, pwd_spaces_start + 3)),
                "bedrooms": text(column(row, pwd_spaces_start + 4)),
            },
            "accessibleEntrance": text(value(row, "Accessible entrance for PwD")),
            "internet": text(value(row, "Is the RAC connected to the internet network?")),
            "hotMeals": text(value(row, "Are people hosted by the Centre offered catered hot meals?")),
            "mealProvider": text(value(row, "Which organisation(s) provide meals?")),
            "foodService": text(value(row, "What kind of food services do you use?")),
            "mealsPerDay": text(value(row, "How many meals per day are being offered to residents?")),
            "cateringSatisfaction": text(value(row, "How satisfied are you with the quality of the catering services provided?")),
            "services": {
                "Psychosocial support (adults)": text(value(row, "Psychological support services (for adults)")),
                "Psychosocial support (children)": text(value(row, "Psychological support services (for children)")),
                "Education": text(value(row, "Education for all ages (including language classes)")),
                "Child protection": text(value(row, "Child protection services")),
                "Information access": text(value(row, "Information access")),
                "Legal advice": text(value(row, "Legal advice")),
                "Health services": text(value(row, "Health services")),
                "Social cohesion": text(value(row, "Social cohesion activities")),
            },
            "needs": {
                "Cooking items": text(value(row, "Centre needs cooking items")),
                "Sleeping items": text(value(row, "Centre needs sleeping items")),
                "Hygiene items": text(value(row, "Centre needs hygiene (personal care) items")),
                "Cleaning items": text(value(row, "Centre needs cleaning items")),
                "Baby & child items": text(value(row, "Centre needs baby and children products")),
                "First aid kit": text(value(row, "Centre needs first aid kit")),
                "Disability supplies": text(value(row, "Centre needs supplies for people with disabilities and children")),
                "Appliances": text(value(row, "Centre needs appliances")),
                "School supplies": text(value(row, "Centre needs school supplies")),
                "Elderly supplies": text(value(row, "Centre needs supplies for ederly people")),
                "Clothes": text(value(row, "Centre needs clothes")),
            },
            "calendar": {
                day: [text(column(row, start + offset)) for offset in range(1, 5)]
                for day, start in calendar_days.items()
            },
        }
        record["demographicProfile"] = kobo_profile(record)
        record["demographicSource"] = "ACTED"
        record["demographicDate"] = record["visitDate"]
        record["mlspDetails"] = None
        records.append(record)
    workbook.close()

    latest_by_rac = {}
    for record in sorted(records, key=lambda item: item["visitDate"] or ""):
        latest_by_rac[record["racId"]] = record

    demographics_records = []
    month_ids = sorted({record["month"] for record in records})
    snapshot_month_ids = sorted(mlsp_snapshots)
    latest_mlsp_capacity = {}
    capacity_by_month = {}
    snapshot_position = 0
    for month in month_ids:
        while snapshot_position < len(snapshot_month_ids) and snapshot_month_ids[snapshot_position] <= month:
            snapshot_month = snapshot_month_ids[snapshot_position]
            for uid, values in mlsp_snapshots[snapshot_month]["rows"].items():
                if values["capacity"] is not None:
                    latest_mlsp_capacity[uid] = values["capacity"]
            snapshot_position += 1
        capacity_by_month[month] = dict(latest_mlsp_capacity)

    for month in month_ids:
        kobo_month = {record["racId"]: record for record in records if record["month"] == month}
        snapshot = mlsp_snapshots.get(month)
        used_ids = set()
        if snapshot:
            for uid, values in snapshot["rows"].items():
                base = dict(kobo_month.get(uid) or latest_by_rac.get(uid) or {})
                base.update(
                    {
                        "racId": uid,
                        "raion": text(base.get("raion")),
                        "city": text(base.get("city")),
                        "address": text(base.get("address")),
                        "female": optional_number(base.get("female")),
                        "male": optional_number(base.get("male")),
                        "otherGender": optional_number(base.get("otherGender")),
                    }
                )
                demographics_records.append(
                    mlsp_demographic_record(
                        base,
                        values,
                        month,
                        snapshot["date"],
                        capacity_by_month[month].get(uid),
                    )
                )
                used_ids.add(uid)
        for uid, record in kobo_month.items():
            if uid not in used_ids:
                fallback = dict(record)
                capacity_uid = CAPACITY_RAC_ALIASES.get(uid, uid)
                fallback["capacity"] = capacity_by_month[month].get(capacity_uid)
                demographics_records.append(fallback)

    months = []
    for month in month_ids:
        month_records = [record for record in records if record["month"] == month]
        month_demographics = [record for record in demographics_records if record["month"] == month]
        dates = [record["visitDate"] for record in month_records if record["visitDate"]]
        mlsp_dates = sorted({record["demographicDate"] for record in month_demographics if record["demographicSource"] == "MLSP"})
        months.append(
            {
                "id": month,
                "from": min(dates),
                "to": max(dates),
                "coverage": len(month_records),
                "demographicsCoverage": len(month_demographics),
                "demographicsFromMLSP": sum(record["demographicSource"] == "MLSP" for record in month_demographics),
                "demographicsDate": mlsp_dates[-1] if mlsp_dates else None,
                "hosted": sum(number(record["hosted"]) for record in month_demographics),
            }
        )

    latest = months[-1]
    previous = months[-2]
    latest_visit = datetime.fromisoformat(latest["to"]).date()
    latest["preliminary"] = latest["coverage"] < previous["coverage"] and (date.today() - latest_visit).days <= 7
    payload = {
        "meta": {
            "source": "ACTED monitoring visits with MLSP demographic precedence",
            "demographicsSource": "Latest monthly MLSP snapshot by RAC ID; ACTED fallback",
            "mapSource": "rac_map.xlsx coordinates matched by RAC ID",
            "reportingRule": "ACTED visits dated on days 1-7 are assigned to the preceding reporting month. MLSP snapshots use calendar months.",
            "latestMonth": latest["id"],
            "asOf": latest["to"],
            "recordCount": len(records),
            "demographicRecordCount": len(demographics_records),
        },
        "months": months,
        "records": records,
        "demographicsRecords": demographics_records,
        "locations": locations,
    }

    TARGET.parent.mkdir(parents=True, exist_ok=True)
    serialized = json.dumps(payload, ensure_ascii=False, separators=(",", ":"), allow_nan=False)
    TARGET.write_text(f"{serialized}\n", encoding="utf-8")
    print(
        f"Wrote {TARGET.relative_to(ROOT)} with {len(records)} ACTED RAC-month records, "
        f"{len(demographics_records)} demographic records and {len(locations)} mapped RACs"
    )


if __name__ == "__main__":
    main()
