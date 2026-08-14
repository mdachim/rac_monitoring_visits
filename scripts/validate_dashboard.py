"""Compact integrity checks for the static dashboard build."""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


def main() -> None:
    data_file = ROOT / "assets" / "data" / "dashboard.json"
    data = json.loads(data_file.read_text(encoding="utf-8"))
    boundary_file = ROOT / "assets" / "data" / "adm1 for PBI with Left Bank.json"
    boundary = json.loads(boundary_file.read_text(encoding="utf-8-sig"))
    records = data["records"]
    demographics = data["demographicsRecords"]

    assert len(records) == 250
    assert len(demographics) == 272
    assert len({(record["month"], record["racId"]) for record in records}) == len(records)
    assert len({(record["month"], record["racId"]) for record in demographics}) == len(demographics)
    assert all(record["capacity"] is not None for record in demographics)
    assert {record["demographicSource"] for record in demographics} <= {"MLSP", "ACTED"}
    assert next(record for record in demographics if record["month"] == "2025-05" and record["racId"] == "573")["capacity"] == 40
    assert len(data["locations"]) == 26
    assert all(45 <= item["latitude"] <= 49 and 26 <= item["longitude"] <= 31 for item in data["locations"])
    assert boundary["type"] == "Topology"
    assert len(boundary["arcs"]) == 135
    assert sum(len(obj.get("geometries", [])) for obj in boundary["objects"].values()) == 37

    june_kobo = [record for record in records if record["month"] == "2026-06"]
    june = [record for record in demographics if record["month"] == "2026-06"]
    july = [record for record in demographics if record["month"] == "2026-07"]
    february = [record for record in demographics if record["month"] == "2026-02"]
    january = [record for record in records if record["month"] == "2026-01"]

    assert len(june_kobo) == 13 and len(june) == 13
    assert sum(record["hosted"] for record in june) == 654
    assert sum(record["capacity"] for record in june) == 871
    assert sum(record["pwd"] for record in june) == 72
    assert {record["demographicDate"] for record in june} == {"2026-06-29"}
    assert all(record["demographicSource"] == "MLSP" for record in june)
    assert len(july) == 11
    assert sum(record["hosted"] for record in july) == 628
    assert sum(record["capacity"] for record in july) == 765
    assert sum(record["pwd"] for record in july) == 74
    assert {record["demographicDate"] for record in july} == {"2026-07-20"}
    assert {record["demographicDate"] for record in february if record["demographicSource"] == "MLSP"} == {"2026-02-23"}
    assert len(january) == 17 and any(record["visitDate"].startswith("2026-02") for record in january)
    assert next(record for record in june if record["racId"] == "2")["raion"] == "Balti"
    assert data["months"][-1]["demographicsDate"] == "2026-07-20"

    expected_sheets = ["Demographics", "Infrastructure", "NFI Needs", "Education & Catering", "Services", "Calendar"]
    download_dir = ROOT / "assets" / "downloads"
    workbook_files = sorted(download_dir.glob("rac-monitoring-visits-*.xlsx"))
    assert len(workbook_files) == len(data["months"])
    for month in data["months"]:
        workbook_file = download_dir / f"rac-monitoring-visits-{month['id']}.xlsx"
        assert workbook_file.exists() and workbook_file.stat().st_size > 12_000
        workbook = load_workbook(workbook_file, read_only=False, data_only=False)
        try:
            assert workbook.sheetnames == expected_sheets
            assert all(f"Reporting month: {month['id']}" in workbook[name]["A2"].value for name in expected_sheets)
            assert workbook["Demographics"]["A4"].value == "RAC ID"
            assert workbook["Infrastructure"]["C4"].value == "Room type"
            assert workbook["Calendar"]["C4"].value == "Monday"
            assert all(workbook[name].freeze_panes == "A5" for name in expected_sheets)
        finally:
            workbook.close()

    html = (ROOT / "index.html").read_text(encoding="utf-8")
    expected_tabs = {
        "Demographics",
        "Infrastructure",
        "NFI Needs",
        "Education &amp; Catering",
        "Services",
        "Calendar",
    }
    assert all(f">{tab}</button>" in html for tab in expected_tabs)
    assert html.index('class="dashboard-overview"') < html.index('class="dashboard-tabs"')
    assert html.index('id="racMap"') < html.index('class="dashboard-tabs"')
    assert html.index('class="dashboard-tabs"') < html.index('id="demographicsView"')
    assert 'role="tablist"' in html and html.count('role="tabpanel"') == 6
    assert 'class="panel-kicker"' not in html
    overview = html[html.index('class="dashboard-overview"'):html.index('class="dashboard-tabs"')]
    demographics_view = html[html.index('id="demographicsView"'):html.index('id="infrastructureView"')]
    assert 'id="ageGenderChart"' in overview and 'id="raionChart"' in overview and 'id="racMap"' in overview
    assert 'id="demographicsTable"' in demographics_view and 'id="racMap"' not in demographics_view
    assert "Collection in progress" not in html
    assert 'src="logos/blue-logo-Moldova.png"' in html
    assert ">Showing<" not in html
    assert 'id="activeFilterBar"' in html
    assert 'id="downloadWorkbook"' in html and "Download XLSX" in html
    assert 'href="assets/downloads/rac-monitoring-visits-2026-07.xlsx"' in html
    assert 'assets/css/styles.css?v=20260814b' in html
    assert 'assets/js/app.js?v=20260814b' in html
    assert 'id="racMap"' in html
    assert "map-placeholder" not in html
    assert "MLSP monthly demographics" in html
    assert "ACTED monitoring visits" in html
    assert "<strong>RAC Monitoring Visits</strong>" in html
    assert "leaflet" not in html.lower()
    assert "assets/js/data.js" not in html
    assert 'class="rank-list raion-scroll"' in html

    css = (ROOT / "assets" / "css" / "styles.css").read_text(encoding="utf-8")
    assert re.search(r"\.bar-track\s*\{[^}]*display:\s*block", css, re.DOTALL)
    assert re.search(r"\.bar-fill\s*\{[^}]*display:\s*block", css, re.DOTALL)
    assert re.search(r"\.rac-map\s*\{[^}]*height:\s*640px", css, re.DOTALL)
    assert re.search(r"\.raion-scroll\s*\{[^}]*height:\s*220px[^}]*overflow-y:\s*auto", css, re.DOTALL)
    assert ".moldova-map" in css and ".moldova-region" in css and ".rac-point" in css
    assert ".map-controls" in css and ".map-help" in css
    assert "touch-action: none" in css
    assert ".timeline-download" in css
    assert re.search(r"\.site-header\s*\{[^}]*margin-bottom:\s*20px", css, re.DOTALL)
    assert re.search(r"h2\s*\{[^}]*text-transform:\s*uppercase", css, re.DOTALL)
    assert re.search(r"\.data-table\s*\{[^}]*border-collapse:\s*separate", css, re.DOTALL)
    assert ".demographic-occupancy-track" in css and ".demographic-age-bar" in css
    assert re.search(r"\.calendar-entry\s*\{[^}]*border-left:\s*2px solid var\(--blue\)", css, re.DOTALL)
    assert "leaflet" not in css.lower()

    app = (ROOT / "assets" / "js" / "app.js").read_text(encoding="utf-8")
    assert "baseDemographicRecords" in app
    assert "renderDemographicTable" in app and "data-demographic-sort" in app
    assert "Female (Kobo)" not in app and "Male (Kobo)" not in app
    assert 'record.demographicSource || "ACTED"' in app
    assert 'fetch("assets/data/dashboard.json")' in app
    assert 'assets/downloads/${workbookName}' in app
    assert 'rac-monitoring-visits-${meta.id}.xlsx' in app
    assert 'fetch("assets/data/adm1%20for%20PBI%20with%20Left%20Bank.json")' in app
    assert "decodeArc" in app and "stitchRing" in app and "racMapPoints" in app
    assert "mapMaxZoom = 8" in app and "zoomMap" in app and 'addEventListener("wheel"' in app
    assert 'addEventListener("pointermove"' in app and "data-base-radius" in app
    assert "13 + Math.sqrt(residents) * 1.1" in app
    assert "window.L" not in app and "tileLayer" not in app
    assert 'rac: (record) => record.racId === key' in app
    show_view = app[app.index("function showView"):app.index("function renderPeriod")]
    assert "activeFilter = null" not in show_view
    assert not (ROOT / "assets" / "js" / "data.js").exists()
    assert not (ROOT / "assets" / "map").exists()

    references = re.findall(r'(?:src|href)="([^"#]+)"', html)
    local_references = [reference for reference in references if not reference.startswith(("http://", "https://"))]
    missing = [reference for reference in local_references if not (ROOT / reference.split("?", 1)[0]).exists()]
    assert not missing, f"Missing local assets: {missing}"

    print(
        json.dumps(
            {
                "koboRecords": len(records),
                "demographicRecords": len(demographics),
                "latestMLSP": {"month": "2026-07", "date": "2026-07-20", "racs": len(july), "hosted": 628},
                "mappedRACs": len(data["locations"]),
                "dataFormat": "JSON",
                "map": "Local Moldova TopoJSON",
                "administrativePolygons": 37,
                "localReferences": "ok",
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
